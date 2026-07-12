"""Tests for the forecast results-export .xlsx surface.

- unit: ``forecast_to_xlsx`` maps a bundle dict into the right sheets.
- integration: ``POST /api/v1/forecast/single.xlsx`` returns a real workbook.
"""
from __future__ import annotations

import io

import openpyxl
from fastapi.testclient import TestClient

from demand_signal_os.api.app import create_app
from demand_signal_os.api.forecast_export import forecast_to_xlsx, leaderboard_to_xlsx

_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_RESULT = {
    "method": "ets",
    "bundle": {
        "sku_id": "WORKBENCH", "location_id": "WORKBENCH",
        "horizon_label": "operational",
        "bucket": {"period": "day"},
        "quantiles": {"q05": 47.8, "q10": 48.7, "q25": 50.2, "q50": 51.8,
                      "q75": 53.4, "q90": 54.9, "q95": 55.7},
        "distribution": {"family": "normal", "params": {"mean": 51.8, "std": 2.4}},
        "mean": 51.8, "fallback_applied": None,
        "provenance": {"forecast_bundle_id": "abc", "model_id": "ets-ZZZ-s7",
                       "seed": 42, "feature_set_hash": "deadbeef",
                       "data_cut_timestamp": "2026-01-01T00:00:00Z",
                       "produced_at": "2026-07-12T00:00:00Z"},
    },
    "band": [
        {"h": 1, "q05": 47.8, "q50": 51.8, "q95": 55.7},
        {"h": 2, "q05": 46.1, "q50": 51.9, "q95": 57.4},
    ],
    "band_truncated": False,
}


def test_forecast_to_xlsx_sheets_and_numeric_cells():
    data = forecast_to_xlsx(_RESULT, history=[50, 52, 48, 55])
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Forecast", "Band", "History", "Provenance"]

    fc = wb["Forecast"]  # first sheet -> title row 1, header row 3
    assert fc.cell(row=1, column=1).value == "DemandSignalOS forecast"
    assert fc.cell(row=3, column=1).value == "Quantile"
    assert fc.cell(row=4, column=1).value == "q05"
    assert isinstance(fc.cell(row=4, column=2).value, (int, float))

    band = wb["Band"]  # non-first -> header row 1
    assert band.cell(row=1, column=1).value == "h"
    assert band.cell(row=2, column=1).value == 1

    hist = wb["History"]
    assert hist.cell(row=2, column=2).value == 50


def test_single_forecast_xlsx_endpoint_returns_workbook():
    with TestClient(create_app()) as client:
        resp = client.post(
            "/api/v1/forecast/single.xlsx",
            json={"history": [10, 12, 9, 11, 13, 10, 12, 11, 10, 13, 9, 12],
                  "horizon": 4, "season_length": 4, "method_id": "ets"},
        )
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"] == _XLSX_CT
        assert "attachment" in resp.headers["content-disposition"]
        assert resp.content[:2] == b"PK"
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        # Forced band=True on the endpoint -> Band sheet always present.
        assert "Forecast" in wb.sheetnames
        assert "Band" in wb.sheetnames
        assert "History" in wb.sheetnames


def test_single_forecast_xlsx_422_on_bad_method():
    with TestClient(create_app()) as client:
        resp = client.post(
            "/api/v1/forecast/single.xlsx",
            json={"history": [1, 2, 3, 4, 5, 6, 7, 8], "method_id": "not_a_method"},
        )
        assert resp.status_code == 422


_LB_RESULT = {
    "config": {"sku_id": "SKU-1", "location_id": "DC-1", "horizon": 4,
               "season_length": 7, "forecaster_set": "all", "n_windows": 2, "seed": 42},
    "entries": [
        {"rank": 1, "method_id": "ets", "is_benchmark": False, "crps": 2.28,
         "smape": 0.11, "pinball_q50": 1.1, "pinball_q90": 0.6, "wis": 2.0,
         "coverage_50": 0.5, "coverage_90": 0.9, "beats_all_benchmarks": True, "n_windows": 2},
        {"rank": 2, "method_id": "seasonal_naive", "is_benchmark": True, "crps": 2.96,
         "smape": None, "pinball_q50": 1.4, "pinball_q90": 0.8, "wis": 2.6,
         "coverage_50": None, "coverage_90": None, "beats_all_benchmarks": None, "n_windows": 2},
    ],
    "winner_method_id": "ets", "winner_is_benchmark": False,
    "feature_set_hash": "abc123", "n_methods": 2, "content_hash": "hash123",
}
_LB_FORECAST = [
    {"h": 1, "q05": 8.1, "q50": 10.0, "q95": 12.2},
    {"h": 2, "q05": 7.6, "q50": 10.1, "q95": 12.9},
    {"h": 3, "q05": 7.0, "q50": 10.2, "q95": 13.5},
    {"h": 4, "q05": 6.5, "q50": 10.3, "q95": 14.1},
]


def test_leaderboard_to_xlsx_includes_ranking_and_forecast():
    data = leaderboard_to_xlsx(_LB_RESULT, _LB_FORECAST)
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Leaderboard", "Winner Forecast", "Summary"]

    lb = wb["Leaderboard"]  # first sheet -> title row 1, header row 3
    assert lb.cell(row=3, column=1).value == "rank"
    assert lb.cell(row=3, column=2).value == "method_id"
    assert lb.cell(row=4, column=2).value == "ets"
    assert isinstance(lb.cell(row=4, column=4).value, (int, float))  # crps numeric

    wf = wb["Winner Forecast"]  # per-horizon forecasted values
    assert wf.cell(row=1, column=1).value == "h"
    assert wf.cell(row=2, column=1).value == 1
    assert wf.cell(row=5, column=1).value == 4  # 4 horizons present
    assert isinstance(wf.cell(row=2, column=2).value, (int, float))

    summ = wb["Summary"]
    fields = [summ.cell(row=r, column=1).value for r in range(1, summ.max_row + 1)]
    assert "winner_method_id" in fields
    assert "content_hash" in fields


def test_leaderboard_to_xlsx_without_forecast_still_ships_ranking():
    data = leaderboard_to_xlsx(_LB_RESULT, None)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Leaderboard", "Summary"]  # no Winner Forecast sheet


def test_single_forecast_yaml_is_simos_arrivals():
    """POST /forecast/single.yaml returns a SimOS-consumable arrivals schedule."""
    import yaml

    with TestClient(create_app()) as client:
        resp = client.post(
            "/api/v1/forecast/single.yaml",
            json={"history": [10, 12, 9, 11, 13, 10, 12, 11, 10, 13, 9, 12],
                  "horizon": 4, "season_length": 4, "method_id": "ets"},
        )
        assert resp.status_code == 200, resp.text
        assert "yaml" in resp.headers["content-type"]
        doc = yaml.safe_load(resp.text)
        sched = doc["sources"][0]["arrivals"]["schedule"]
        assert len(sched) == 4  # one entry per horizon step
        assert sched[0]["time"] == 0.0
        assert "rate_per_hour" in sched[0]
