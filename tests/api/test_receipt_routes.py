"""Trust-gate receipt endpoint tests for the thin DSO API (DECISIONS_LOG §P #65).

Mirrors the SimOS reference test
(``simulation_os/tests/unit/api/test_calibration_receipt_routes.py``). Guarded so
it runs where trust_gate + excel_io + fastapi are installed (CI / the API image)
and skips cleanly elsewhere — and never pulls the heavy DSO forecasting stack.
"""

from __future__ import annotations

import io

import pytest

pytest.importorskip("trust_gate")
pytest.importorskip("excel_io")
pytest.importorskip("fastapi")

import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from ops_schemas import (  # noqa: E402
    CalibrationReceipt,
    derive_test_mode_secret,
    verify_receipt,
)

from demand_signal_os.api.app import create_app  # noqa: E402


@pytest.fixture()
def client() -> TestClient:
    return TestClient(create_app())


def test_health(client: TestClient) -> None:
    resp = client.get("/api/v1/health")
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"status": "ok", "engine": "demandsignal"}


def test_emit_receipt_is_signed_and_verifies(client: TestClient) -> None:
    resp = client.post(
        "/api/v1/calibration/receipt",
        json={
            "sku_id": "SKU-9000",
            "location_id": "DC-WEST",
            "horizon_label": "operational",
            "baseline_crps": 10.0,
            "actual_count": 21,
            "checks": [
                {
                    "name": "90% interval coverage",
                    "measured_value": 0.91,
                    "reference_value": 0.90,
                    "tolerance": 0.05,
                    "direction": "match",
                },
                {
                    "name": "CRPS vs baseline (ratio)",
                    "measured_value": 0.86,
                    "reference_value": 1.0,
                    "tolerance": 0.5,
                    "direction": "lower_better",
                },
            ],
        },
    )
    assert resp.status_code == 200, resp.text
    receipt = CalibrationReceipt.model_validate(resp.json())
    assert verify_receipt(receipt, derive_test_mode_secret("demandsignal"))
    assert all(ph.passed for ph in receipt.phases)


def test_example_receipt_endpoint(client: TestClient) -> None:
    resp = client.get("/api/v1/calibration/receipt/example")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["provenance"]["engine"] == "demandsignal"
    assert len(body["phases"][0]["metrics"]) == 3
    assert body["signature"]
    # The signed example must actually verify under the engine's test-mode secret.
    receipt = CalibrationReceipt.model_validate(body)
    assert verify_receipt(receipt, derive_test_mode_secret("demandsignal"))


def test_receipt_xlsx_export(client: TestClient) -> None:
    receipt = client.get("/api/v1/calibration/receipt/example").json()
    resp = client.post("/api/v1/calibration/receipt/xlsx", json=receipt)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Verdict" in wb.sheetnames and "Checks" in wb.sheetnames
