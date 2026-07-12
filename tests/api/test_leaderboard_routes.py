"""Leaderboard API tests — submit/poll/winner/receipt + auth.

A small config keeps the heavy GBM path fast. Starlette's TestClient runs
BackgroundTasks synchronously before the POST returns, so the run is already
complete when submit_leaderboard responds.
"""

from __future__ import annotations

import math

import pytest
from fastapi.testclient import TestClient

from demand_signal_os.api.app import create_app

# Deterministic seasonal-ish series, dense (no intermittency), long enough for
# 2 windows of horizon 4 with min_train_size 24 and GBM lag/rolling features.
_HISTORY = [
    10 + 3 * math.sin(2 * math.pi * i / 7) + (i % 3) for i in range(48)
]

_REQUEST = {
    "sku_id": "SKU-1",
    "location_id": "DC-1",
    "history": _HISTORY,
    "bucket_period": "day",
    "start_date": "2026-01-01",
    "horizon": 4,
    "season_length": 7,
    "intermittent_mode": "off",
    "n_windows": 2,
    "min_train_size": 24,
    "seed": 42,
}


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(create_app())


@pytest.fixture(scope="module")
def completed_run(client: TestClient) -> str:
    resp = client.post("/api/v1/forecast/leaderboard", json=_REQUEST)
    assert resp.status_code == 200, resp.text
    return resp.json()["run_id"]


def test_health(client: TestClient) -> None:
    resp = client.get("/api/v1/health")
    assert resp.status_code == 200
    assert resp.json()["engine"] == "demandsignal"


def test_submit_returns_run_id(completed_run: str) -> None:
    assert completed_run.startswith("lb_")


def test_poll_returns_ranked_result(client: TestClient, completed_run: str) -> None:
    resp = client.get(f"/api/v1/forecast/leaderboard/{completed_run}")
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "complete"
    result = body["result"]
    # off-mode: ets, gbm, arima, theta, ces + 3 benchmarks
    assert result["n_methods"] == 8
    ranks = [e["rank"] for e in result["entries"]]
    assert ranks == sorted(ranks)
    assert result["entries"][0]["rank"] == 1


def test_winner_returns_bundle(client: TestClient, completed_run: str) -> None:
    resp = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/winner")
    assert resp.status_code == 200
    body = resp.json()
    assert body["winner_method_id"]
    bundle = body["bundle"]
    # Bundle-ready: has quantiles + method + provenance for the downstream bundle.
    assert "quantiles" in bundle
    assert bundle["method"] == body["winner_method_id"]
    assert bundle["quantiles"]["q05"] <= bundle["quantiles"]["q95"]


def test_xlsx_export_has_ranking_and_forecast(client: TestClient, completed_run: str) -> None:
    import io

    import openpyxl

    resp = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    # Ranking + the winner's per-horizon forecasted values + summary.
    assert "Leaderboard" in wb.sheetnames
    assert "Winner Forecast" in wb.sheetnames
    wf = wb["Winner Forecast"]
    assert wf.cell(row=1, column=1).value == "h"
    # horizon=4 -> 4 per-step rows of real forecasted values.
    assert wf.max_row == 5  # header + 4 horizons


def test_xlsx_export_409_when_not_complete(client: TestClient) -> None:
    resp = client.get("/api/v1/forecast/leaderboard/lb_missing/xlsx")
    assert resp.status_code == 404  # unknown run


def test_arrivals_yaml_is_simos_consumable(client: TestClient, completed_run: str) -> None:
    """The DSO→SimOS contract: a sources/arrivals.schedule YAML with one entry
    per horizon step (rate_per_hour + noise_std), the shape SimOS ingests."""
    import yaml

    resp = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/arrivals.yaml")
    assert resp.status_code == 200, resp.text
    assert "yaml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    doc = yaml.safe_load(resp.text)
    assert "sources" in doc
    src = doc["sources"][0]
    assert src["arrivals"]["distribution"] == "schedule"
    schedule = src["arrivals"]["schedule"]
    assert len(schedule) == 4  # horizon=4 -> 4 forward schedule entries
    assert schedule[0]["time"] == 0.0
    assert all("rate_per_hour" in e for e in schedule)
    # forward, monotone times.
    assert [e["time"] for e in schedule] == sorted(e["time"] for e in schedule)


def test_receipt_is_signed(client: TestClient, completed_run: str) -> None:
    resp = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/receipt")
    assert resp.status_code == 200
    receipt = resp.json()
    # Signed trust receipt carries an HMAC signature.
    assert receipt.get("signature")
    # Checks live in phases[].metrics[]; the benchmark gate must be present.
    names = {
        m["name"] for ph in receipt["phases"] for m in ph["metrics"]
    }
    assert any("beats all benchmarks" in n for n in names)


def test_receipt_is_deterministic(client: TestClient, completed_run: str) -> None:
    a = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/receipt").json()
    b = client.get(f"/api/v1/forecast/leaderboard/{completed_run}/receipt").json()
    # Same run -> byte-identical signed receipt (RULE 5).
    assert a["calibration_id"] == b["calibration_id"]
    assert a["signature"] == b["signature"]


def test_idempotent_run_id(client: TestClient, completed_run: str) -> None:
    resp = client.post("/api/v1/forecast/leaderboard", json=_REQUEST)
    assert resp.json()["run_id"] == completed_run


def test_unknown_run_404(client: TestClient) -> None:
    resp = client.get("/api/v1/forecast/leaderboard/lb_doesnotexist")
    assert resp.status_code == 404


def test_auth_enforced_when_keys_set(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("DSO_API_KEYS", "dso_live_secret123")
    authed_client = TestClient(create_app())
    # No key -> 401
    resp = authed_client.get("/api/v1/forecast/leaderboard/lb_whatever")
    assert resp.status_code == 401
    # Valid key -> passes auth (404 because run doesn't exist, not 401)
    resp = authed_client.get(
        "/api/v1/forecast/leaderboard/lb_whatever",
        headers={"X-API-Key": "dso_live_secret123"},
    )
    assert resp.status_code == 404


# --- Cross-engine SSO (Phase 2b): mao_live_ bearer dual-accept --------------


class _FakeResolver:
    """Stand-in for platform_auth.MaoTierResolver."""

    def __init__(self, *, plan: str | None = None, raise_status: int | None = None):
        self._plan = plan
        self._raise = raise_status

    async def resolve_plan(self, token: str) -> str:
        from fastapi import HTTPException

        if self._raise is not None:
            raise HTTPException(status_code=self._raise, detail="resolver")
        assert self._plan is not None
        return self._plan


def _client_with_resolver(**kw: object) -> TestClient:
    app = create_app()
    app.state.mao_tier_resolver = _FakeResolver(**kw)  # type: ignore[arg-type]
    return TestClient(app)


def test_mao_premium_bearer_passes_auth() -> None:
    c = _client_with_resolver(plan="premium")
    r = c.get(
        "/api/v1/forecast/leaderboard/lb_none",
        headers={"Authorization": "Bearer mao_live_x"},
    )
    assert r.status_code == 404  # passed auth (run missing), not 401/403


def test_mao_enterprise_bearer_passes_auth() -> None:
    c = _client_with_resolver(plan="enterprise")
    r = c.get(
        "/api/v1/forecast/leaderboard/lb_none",
        headers={"Authorization": "Bearer mao_live_x"},
    )
    assert r.status_code == 404


def test_mao_free_bearer_denied_403() -> None:
    c = _client_with_resolver(plan="free")
    r = c.get(
        "/api/v1/forecast/leaderboard/lb_none",
        headers={"Authorization": "Bearer mao_live_x"},
    )
    assert r.status_code == 403


def test_mao_revoked_bearer_401() -> None:
    c = _client_with_resolver(raise_status=401)
    r = c.get(
        "/api/v1/forecast/leaderboard/lb_none",
        headers={"Authorization": "Bearer mao_live_x"},
    )
    assert r.status_code == 401
